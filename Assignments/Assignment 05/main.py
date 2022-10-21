# Python imports
import openpyxl
import requests
import json

# Packages imports
from openpyxl.chart import BarChart, Reference


OPEN_WEATHER_MAP_API_KEY = "5111c4dcad126d1b6612b42b6a34f171"
EXCEL_FILE_NAME = "CitiesTemperature.xlsx"


def weather_city(city):
    """
    This function will return the weather of the city
    :param city:
    :return:
    """
    # URL f-string
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid={OPEN_WEATHER_MAP_API_KEY}'

    # get the response from the API
    response = requests.get(url).text

    print(f"Today in {city.title()} the temperature is {round(json.loads(response)['main']['temp'] - 273.15, 2)}°C")
    return round(json.loads(response)['main']['temp'] - 273.15, 2)


_cities = ['Karachi', 'Sudbury', 'Lahore', 'Islamabad', 'Multan']
_temperatures = {city: weather_city(city=city) for city in _cities}


def create_workbook_object():
    """
    This function will create and return workbook object
    :return:
    """
    wb = openpyxl.Workbook()
    return wb.active


def filling_data(ws):
    """
    This function will fill data into the object of workbook that has been created earlier.
    :param ws:
    :return:
    """
    ws["A1"] = "City"
    ws["B1"] = "Temperature"

    for i in range(1, len(_cities) + 1):
        ws[f"A{str(i + 1)}"] = _cities[i - 1]
        ws[f"B{str(i + 1)}"] = _temperatures[_cities[i - 1]]


def create_chart_within_workbook(ws):
    """
    This function will create barchart using the workbook reference data
    :param ws:
    :return:
    """
    values = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=6)
    titles = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Barchart for temperature of different cities"
    chart.y_axis.title = 'Temperature'
    chart.x_axis.title = 'City'
    chart.add_data(values)
    chart.set_categories(titles)
    ws.add_chart(chart, "D1")


def save_workbook_data(ws):
    """
    This function will save the file into the same directory as the main file.
    :param ws:
    :return:
    """
    ws.save(f"{EXCEL_FILE_NAME}")


def _run():
    """
    This is an abstract layer which will handle the execution process
    :return:
    """
    workbook_object = create_workbook_object()
    filling_data(workbook_object)


if __name__ == '__main__':
    _run()
